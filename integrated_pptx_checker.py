"""
통합 PPTX 검사 파이프라인

3단계 검사:
1. PPTX 텍스트 추출
2. 파인튜닝 모델 + pptx_checker 규칙 기반 검사 (합집합)
3. Claude API로 최종 검토 (문맥 기반)

사용법:
    python integrated_pptx_checker.py input.pptx
    python integrated_pptx_checker.py input.pptx --output report.json
"""

import json
import os
import re
from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET
from typing import List, Dict, Tuple
from collections import defaultdict
import difflib

import torch
from peft import AutoPeftModelForCausalLM
from transformers import AutoTokenizer

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

# pptx_checker 모듈 임포트
try:
    from pptx_checker import PPTXChecker, check_with_claude
except ImportError:
    print("⚠️ pptx_checker.py를 찾을 수 없습니다.")
    PPTXChecker = None
    check_with_claude = None


# ============================================
# Stage 1: PPTX 텍스트 추출 (Ultra Think 모드)
# ============================================

def extract_texts_from_pptx(pptx_path: Path) -> List[Dict]:
    """
    Ultra Think 모드로 PPTX에서 텍스트 추출
    XML 추출 + Claude API 구조 분류 (빠른 속도, 높은 정확도)
    """
    print("\n" + "=" * 70)
    print("🚀 Stage 1: Ultra Think 텍스트 추출 (XML + Claude API)")
    print("=" * 70)

    try:
        from ultra_text_extractor import ultra_extract_texts

        # Ultra Think 추출 실행 (XML + Claude API 방식)
        result = ultra_extract_texts(str(pptx_path))

        # 포맷 변환 (기존 integrated_pptx_checker 형식으로)
        texts = []
        for slide_data in result['slides']:
            # 구조화된 필드들을 텍스트로 추가
            for field_name in ['chapter', 'title', 'subtitle', 'lead']:
                field_value = slide_data.get(field_name, '')
                if field_value and len(field_value) > 2:
                    texts.append({
                        'slide': slide_data['page'],
                        'text': field_value,
                        'char_count': len(field_value),
                        'field': field_name
                    })

            # Contents 추가
            for text in slide_data.get('contents', []):
                if len(text) > 2:  # 최소 길이 필터
                    texts.append({
                        'slide': slide_data['page'],
                        'text': text,
                        'char_count': len(text),
                        'field': 'contents'
                    })

        print(f"\n✅ Ultra Think 추출 완료: {len(texts)}개 텍스트 블록")
        print(f"   ({result['total_slides']}개 슬라이드, 구조 분류 완료)")

        return texts

    except ImportError as e:
        print(f"⚠️  Ultra Think 모듈 로드 실패, 기본 XML 추출로 전환: {e}")
        # Fallback to basic XML extraction
        return _extract_texts_xml_only(pptx_path)
    except Exception as e:
        print(f"❌ Ultra Think 추출 실패, 기본 XML 추출로 전환: {e}")
        import traceback
        traceback.print_exc()
        return _extract_texts_xml_only(pptx_path)


def _extract_texts_xml_only(pptx_path: Path) -> List[Dict]:
    """기본 XML 기반 텍스트 추출 (fallback)"""
    print("\n📋 Fallback: XML 기반 텍스트 추출")

    texts = []

    try:
        with ZipFile(pptx_path, 'r') as zf:
            slide_files = sorted([
                f for f in zf.namelist()
                if f.startswith('ppt/slides/slide') and f.endswith('.xml')
            ], key=lambda x: int(re.search(r'slide(\d+)', x).group(1)))

            for slide_idx, slide_file in enumerate(slide_files):
                slide_num = slide_idx + 1

                with zf.open(slide_file) as f:
                    tree = ET.parse(f)
                    root = tree.getroot()

                    ns = {
                        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                        'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
                    }

                    # 각 paragraph별로 텍스트 추출
                    for sp in root.findall('.//p:sp', ns):
                        txBody = sp.find('.//p:txBody', ns)
                        if txBody is None:
                            continue

                        for para in txBody.findall('a:p', ns):
                            para_text = []
                            for t_elem in para.findall('.//a:t', ns):
                                if t_elem.text:
                                    para_text.append(t_elem.text)

                            text = ''.join(para_text).strip()
                            if text and len(text) > 5:
                                texts.append({
                                    'slide': slide_num,
                                    'text': text,
                                    'char_count': len(text)
                                })

    except Exception as e:
        print(f"❌ XML 추출 실패: {e}")
        return []

    print(f"✅ XML 추출 완료: {len(texts)}개 텍스트 블록")
    return texts


# ============================================
# Stage 2: 파인튜닝 모델 + 규칙 기반 검사
# ============================================

def check_with_finetuned_model(texts: List[Dict], model_path: str) -> List[Dict]:
    """파인튜닝된 모델로 검사 (전수 검사)"""
    print("\n📝 파인튜닝 모델 검사 중...")

    # 완전 기본 설정 (기존 동작하던 방식)
    device = "mps" if torch.backends.mps.is_available() else "cpu"

    try:
        # device_map 제거, 기본 방식으로 로드
        model = AutoPeftModelForCausalLM.from_pretrained(model_path)
        model.to(device)
        model.eval()

        tokenizer = AutoTokenizer.from_pretrained(model_path)
        print(f"   모델 로드 완료 (디바이스: {device})")

    except Exception as e:
        print(f"   ⚠️ 모델 로드 실패: {e}")
        return []

    issues = []

    # 범위 확대 (5-150자)
    filtered_texts = [t for t in texts if 5 <= t['char_count'] <= 150]
    print(f"   검사 대상: {len(filtered_texts)}개 (5-150자)")
    print(f"   🔥 전수 검사 모드: 모든 텍스트를 검사합니다")

    for idx, item in enumerate(filtered_texts, 1):
        text = item['text']
        slide_num = item['slide']

        try:
            # 프롬프트 생성
            prompt = f"다음 문장의 맞춤법과 용어 통일성을 검사해주세요. 문제가 없으면 '문제 없음'이라고만 답하세요:\n\n{text}"

            messages = [{"role": "user", "content": prompt}]

            text_input = tokenizer.apply_chat_template(
                messages,
                tokenize=False,
                add_generation_prompt=True
            )

            inputs = tokenizer(text_input, return_tensors="pt").to(device)

            with torch.no_grad():
                outputs = model.generate(
                    **inputs,
                    max_new_tokens=150,
                    temperature=0.3,  # 더 많은 이슈 감지를 위해 0.1 → 0.3으로 조정
                    do_sample=True,
                    pad_token_id=tokenizer.pad_token_id,
                    eos_token_id=tokenizer.eos_token_id,
                )

            response = tokenizer.decode(outputs[0], skip_special_tokens=True)

            # 응답 파싱
            if "assistant" in response:
                response = response.split("assistant")[-1].strip()

            response = re.sub(r'<think>.*?</think>', '', response, flags=re.DOTALL).strip()

            # 문제가 있는지 확인
            if "수정된 문장" in response and "문제 없음" not in response:
                match = re.search(r'수정된 문장[:\s]*\n([^\n]+)', response)
                if match:
                    corrected = match.group(1).strip()

                    # 실제 변경사항이 있는지 확인
                    if corrected != text:
                        issue_type = "맞춤법"
                        if "띄어쓰기" in response:
                            issue_type = "띄어쓰기"
                        elif "용어통일" in response:
                            issue_type = "용어통일"

                        issues.append({
                            'source': 'finetuned_model',
                            'slide': slide_num,
                            'original': text,
                            'suggested': corrected,
                            'type': issue_type,
                            'confidence': 'medium'
                        })
                        print(f"   [{idx}/{len(filtered_texts)}] 슬라이드 {slide_num} - ⚠️ 발견")
                    else:
                        print(f"   [{idx}/{len(filtered_texts)}] 슬라이드 {slide_num} - ✅", end='\r')
                else:
                    print(f"   [{idx}/{len(filtered_texts)}] 슬라이드 {slide_num} - ✅", end='\r')
            else:
                print(f"   [{idx}/{len(filtered_texts)}] 슬라이드 {slide_num} - ✅", end='\r')

        except Exception as e:
            print(f"\n   ❌ 검사 실패 (슬라이드 {slide_num}): {e}")

    print(f"\n   발견된 문제: {len(issues)}건")
    return issues


def check_with_rules(pptx_path: Path) -> List[Dict]:
    """규칙 기반 검사 (pptx_checker)"""
    print("\n📋 규칙 기반 검사 중...")

    if PPTXChecker is None:
        print("   ⚠️ pptx_checker를 사용할 수 없습니다")
        return []

    try:
        checker = PPTXChecker(str(pptx_path))
        checker.extract_text()
        checker.check_term_consistency()
        checker.check_typos()
        checker.check_number_format()

        # 이슈를 통합 포맷으로 변환
        issues = []
        for issue in checker.issues:
            if issue['type'] == '용어 불일치':
                issues.append({
                    'source': 'rule_based',
                    'slide': None,  # 여러 슬라이드에 걸쳐 있을 수 있음
                    'original': ', '.join(issue['found']),
                    'suggested': issue['recommended'],
                    'type': '용어통일',
                    'confidence': 'high',
                    'locations': issue.get('locations', [])
                })
            elif issue['type'] == '맞춤법 오류':
                issues.append({
                    'source': 'rule_based',
                    'slide': issue['slide'],
                    'original': issue['wrong'],
                    'suggested': issue['correct'],
                    'type': '맞춤법',
                    'confidence': 'high',
                    'context': issue.get('context', '')
                })
            elif issue['type'] == '띄어쓰기 오류':
                issues.append({
                    'source': 'rule_based',
                    'slide': issue['slide'],
                    'original': issue['pattern'],
                    'suggested': issue['correct'],
                    'type': '띄어쓰기',
                    'confidence': 'high',
                    'context': issue.get('context', '')
                })

        print(f"   발견된 문제: {len(issues)}건")
        return issues

    except Exception as e:
        print(f"   ❌ 규칙 기반 검사 실패: {e}")
        return []


def merge_issues(model_issues: List[Dict], rule_issues: List[Dict]) -> List[Dict]:
    """파인튜닝 모델과 규칙 기반 검사 결과 합치기 (합집합)"""
    print("\n🔄 검사 결과 통합 중...")

    # 중복 제거를 위한 키 생성
    seen = set()
    merged = []

    # 규칙 기반은 신뢰도가 높으므로 우선 추가
    for issue in rule_issues:
        key = (issue.get('slide'), issue['original'], issue['suggested'])
        if key not in seen:
            seen.add(key)
            merged.append(issue)

    # 모델 결과 추가 (중복 제거)
    for issue in model_issues:
        key = (issue.get('slide'), issue['original'], issue['suggested'])
        if key not in seen:
            seen.add(key)
            merged.append(issue)

    print(f"   통합 결과: {len(merged)}건 (모델: {len(model_issues)}, 규칙: {len(rule_issues)})")
    return merged


# ============================================
# Stage 3: Claude API 최종 검토
# ============================================

def load_fewshot_examples():
    """Few-shot 예시 로드 (Knowledge Distillation)"""
    try:
        fewshot_path = Path(__file__).parent / 'fewshot_examples.json'
        with open(fewshot_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"   ⚠️  Few-shot 예시 로드 실패: {e}")
        return {"positive_examples": [], "negative_examples": []}


def review_with_claude(pptx_path: Path, texts: List[Dict], merged_issues: List[Dict]) -> Dict:
    """Claude API로 최종 검토 및 문맥 기반 검사"""
    print("\n🤖 Stage 3: Claude API 최종 검토")
    print("=" * 70)

    if check_with_claude is None:
        print("   ⚠️ Claude API를 사용할 수 없습니다")
        return {}

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("   ⚠️ ANTHROPIC_API_KEY가 설정되지 않았습니다")
        print("   Claude API 검토를 건너뜁니다")
        return {}

    # Claude에게 기존 발견사항을 전달
    existing_issues_text = "\n".join([
        f"- 슬라이드 {issue.get('slide', '?')}: \"{issue['original']}\" → \"{issue['suggested']}\" ({issue['type']})"
        for issue in merged_issues[:20]  # 최대 20개만
    ])

    # 텍스트 타입별로 그룹핑 (Chapter, Title, Subtitle, Lead, Contents 분리)
    print(f"   🔥 타입별 그룹 검사 모드: {len(texts)}개 텍스트를 타입별로 분류합니다")

    texts_by_type = {
        'chapter': [],
        'title': [],
        'subtitle': [],
        'lead': [],
        'contents': []
    }

    for t in texts:
        text_type = t.get('field', 'contents').lower()
        if text_type in texts_by_type:
            texts_by_type[text_type].append(t)
        else:
            texts_by_type['contents'].append(t)

    # 타입별 개수 출력
    print(f"   📊 타입별 분류:")
    for type_name, type_texts in texts_by_type.items():
        if len(type_texts) > 0:
            print(f"      • {type_name}: {len(type_texts)}개")

    # 결과 누적 변수 초기화
    all_context_issues = []
    all_false_positives = []
    all_additional_issues = []

    # Few-shot 예시 로드 (Knowledge Distillation)
    fewshot = load_fewshot_examples()
    positive_examples = fewshot.get('positive_examples', [])
    negative_examples = fewshot.get('negative_examples', [])

    # Positive 예시 포맷팅 (상위 25개)
    positive_text = "\n".join([
        f"{i+1}. \"{ex['wrong']}\" → \"{ex['correct']}\" ({ex['reason']})"
        for i, ex in enumerate(positive_examples[:25])
    ])

    # Negative 예시 포맷팅 (전체)
    negative_text = "\n".join([
        f"{i+1}. \"{ex['text']}\" → 변경 금지 ({ex['reason']})"
        for i, ex in enumerate(negative_examples)
    ])

    print(f"   📚 Few-shot 예시 로드: {len(positive_examples)}개 positive, {len(negative_examples)}개 negative")

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        # 타입별로 순차 처리
        type_order = ['chapter', 'title', 'subtitle', 'lead', 'contents']

        for type_name in type_order:
            type_texts = texts_by_type[type_name]
            if len(type_texts) == 0:
                continue

            print(f"\n   📝 [{type_name.upper()}] 검사 중 ({len(type_texts)}개)...")

            # 타입별로 배치 처리
            batch_size = 200
            num_batches = (len(type_texts) + batch_size - 1) // batch_size

            for batch_idx in range(num_batches):
                start_idx = batch_idx * batch_size
                end_idx = min((batch_idx + 1) * batch_size, len(type_texts))
                batch_texts = type_texts[start_idx:end_idx]

                print(f"      배치 {batch_idx + 1}/{num_batches} 처리 중 ({len(batch_texts)}개 텍스트)...")

                batch_text = "\n".join([f"[슬라이드 {t['slide']}] {t['text']}" for t in batch_texts])

                prompt = f"""다음은 H-온드림 임팩트 스타트업 지원 프로그램 PPT 문서입니다.

이번 배치는 **{type_name.upper()}** 타입 텍스트입니다.
- CHAPTER: 슬라이드 섹션 제목 (예: "01. 발견", "02. 채용")
- TITLE: 슬라이드 주 제목
- SUBTITLE: 슬라이드 부제목
- LEAD: 리드문 (슬라이드 설명문)
- CONTENTS: 본문 내용

**{type_name.upper()} 타입 특화 검사:**
{'- 섹션 번호 형식 일관성 (01., 02., ...)' if type_name == 'chapter' else ''}
{'- 제목 길이 및 톤 일관성' if type_name == 'title' else ''}
{'- 부제목 스타일 통일' if type_name == 'subtitle' else ''}
{'- 리드문 길이 및 흐름 일관성 (25-35자 권장)' if type_name == 'lead' else ''}
{'- 본문 내용 용어 통일' if type_name == 'contents' else ''}

**📚 학습 예시 (이런 패턴을 찾아주세요):**

✅ 올바른 교정 예시:
{positive_text}

❌ 변경하면 안 되는 예시:
{negative_text}

**✅ 반드시 변경해야 할 것 (필수 규칙):**
1. 브랜드명: "H온드림", "H 온드림", "H-OnDream" → "H-온드림"
2. 용어 통일:
   - "펠로우" → "펠로" (모든 경우)
   - "기업가정신", "기업가 정신" → "임팩트 앙트프러너십"
   - "impact", "Impact" → "임팩트"
   - "Mentoring" → "멘토링"
   - "펜테크" → "핀테크"
   - "엑셀러레이터" → "액셀러레이터"
   - "Programme" → "프로그램"
   - "소셜 벤처" → "소셜벤처" (붙여쓰기)

3. 띄어쓰기 오류:
   - "될수록", "할수있", "될수있" → "될 수록", "할 수 있", "될 수 있"
   - "~통한", "~관한", "~따른" → "~을 통한", "~에 관한", "~에 따른"
   - "두번째", "첫번째" → "두 번째", "첫 번째"
   - "억원" → "억 원" (단, 이미 "억 원"으로 띄어져 있으면 변경하지 말 것!)

4. 조사 중복 제거:
   - "~와의 ~될" → "~와 ~될"
   - "~로의 ~되는" → "~로 ~되는"
   - "실무와의 연결될" → "실무와 연결될"

5. 대소문자 통일:
   - "Ai" → "AI" (무조건)
   - "mysclass", "MYSCLASS" → "MYSClass"

6. 맞춤법:
   - "로써" → 문맥 보고 "로서" 판단 (자격/지위: "로서", 수단/방법: "로써")

**🔍 문맥 판단 및 스타일 개선 (additional_issues로 제안):**

1. "스타트업" → 임팩트/소셜 맥락이면 "임팩트 스타트업" 제안, 일반 생태계는 유지

2. **의미 개선 제안**:
   - "정성 성과" → "장기적 성장성과", "질적 성과" 제안 (더 명확한 표현)
   - "연구" → 부담스럽다면 "사례 발굴", "특성 정립", "분석" 제안
   - 모호한 표현 → 구체적 표현 제안

3. **스타일 개선**:
   - 제목이 평범하면 → 더 강조하거나 구체화 제안
   - 리드문 흐름 개선 제안
   - 전문 용어 vs 쉬운 표현 균형 제안

4. **일관성 검토**:
   - 같은 개념을 다르게 표현 → 통일 제안
   - 연도 순서 검증 (2024 → 2025 → 2026)
   - 숫자 표기 일관성

**⚠️ 중요: False Positive 주의사항**
- Executive, Summary, Insight, Station F, Co-creation 같은 영어는 변경하지 마세요
- "길잡이", "올바른" 같은 정상 한국어를 틀린 단어로 바꾸지 마세요
- "AI shift", "Keyword" 같은 일반 영어 용어도 변경하지 마세요

**이미 발견된 문제들 (참고용):**
{existing_issues_text if existing_issues_text else "없음"}

**텍스트:**
{batch_text}

JSON 형식으로 응답:
{{
  "context_based_issues": [
    {{"slide": 번호, "original": "원본", "suggested": "제안", "reason": "이유", "confidence": "high/medium/low"}}
  ],
  "false_positives": [
    {{"slide": 번호, "original": "원본", "reason": "왜 false positive인지"}}
  ],
  "additional_issues": [
    {{"slide": 번호, "issue": "문제", "suggestion": "제안", "reason": "이유"}}
  ]
}}"""

            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4000,
                messages=[{"role": "user", "content": prompt}]
            )

            # JSON 파싱
            content = response.content[0].text
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0].strip()
            elif "```" in content:
                content = content.split("```")[1].split("```")[0].strip()

            batch_result = json.loads(content)

            # 결과 누적
            batch_context_count = len(batch_result.get('context_based_issues', []))
            batch_fp_count = len(batch_result.get('false_positives', []))
            batch_additional_count = len(batch_result.get('additional_issues', []))

            if batch_result.get('context_based_issues'):
                all_context_issues.extend(batch_result['context_based_issues'])
            if batch_result.get('false_positives'):
                all_false_positives.extend(batch_result['false_positives'])
            if batch_result.get('additional_issues'):
                all_additional_issues.extend(batch_result['additional_issues'])

            print(f"      ✓ 배치 완료 - 문맥 이슈: {batch_context_count}건, FP: {batch_fp_count}건, 추가: {batch_additional_count}건")

        # 전체 결과 집계
        result = {
            'context_based_issues': all_context_issues,
            'false_positives': all_false_positives,
            'additional_issues': all_additional_issues
        }

        print(f"\n   ✅ Claude 전수 검토 완료!")
        print(f"      문맥 기반 이슈: {len(all_context_issues)}건")
        print(f"      False positives: {len(all_false_positives)}건")
        print(f"      추가 이슈: {len(all_additional_issues)}건")

        return result

    except Exception as e:
        print(f"   ❌ Claude API 오류: {e}")
        import traceback
        traceback.print_exc()
        return {
            'context_based_issues': all_context_issues,
            'false_positives': all_false_positives,
            'additional_issues': all_additional_issues
        }


# ============================================
# 파인튜닝 이유 생성 & Excel 출력
# ============================================

def generate_reasons_for_finetuned(issues: list) -> list:
    """
    파인튜닝 모델 결과에 대해 Claude API로 이유 생성

    Args:
        issues: 이슈 리스트

    Returns:
        이유가 추가된 이슈 리스트
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("⚠️  ANTHROPIC_API_KEY 없음 - 파인튜닝 이유 생성 스킵")
        return issues

    # 파인튜닝 이슈만 필터
    finetuned_issues = [
        issue for issue in issues
        if issue.get('source') == 'finetuned_model' and not issue.get('reason')
    ]

    if not finetuned_issues:
        return issues

    print(f"\n🤖 Claude API로 파인튜닝 결과 이유 생성 중... ({len(finetuned_issues)}건)")

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        # 배치로 처리
        batch_data = []
        for issue in finetuned_issues:
            batch_data.append({
                "original": issue.get('original', ''),
                "suggested": issue.get('suggested', ''),
                "type": issue.get('type', '')
            })

        prompt = f"""다음은 한국어 맞춤법/용어 통일성 검사기가 발견한 수정 제안입니다.
각 수정 제안에 대해 **왜 수정이 필요한지** 간결하게 설명해주세요 (10-20자).

입력:
```json
{json.dumps(batch_data, ensure_ascii=False, indent=2)}
```

출력 형식 (JSON만):
```json
{{
  "reasons": [
    "설명1",
    "설명2",
    ...
  ]
}}
```

JSON만 출력하세요."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}]
        )

        response_text = message.content[0].text

        # JSON 추출
        json_str = response_text.strip()
        if '```json' in json_str:
            json_str = json_str.split('```json')[1].split('```')[0].strip()
        elif '```' in json_str:
            json_str = json_str.split('```')[1].split('```')[0].strip()

        data = json.loads(json_str)
        reasons = data.get('reasons', [])

        # 이슈에 이유 추가
        issue_map = {id(issue): issue for issue in issues}
        for i, finetuned_issue in enumerate(finetuned_issues):
            if i < len(reasons):
                issue_map[id(finetuned_issue)]['reason'] = reasons[i]

        print(f"   ✅ 이유 생성 완료")

    except Exception as e:
        print(f"   ⚠️  이유 생성 실패: {e}")

    return list(issue_map.values())


def save_report_as_excel(report: dict, excel_path: str, all_texts: List[Dict] = None):
    """리포트를 Excel 파일로 저장 (변경 부분만 빨간색 표시 + 문제 없음 항목 포함)"""
    from openpyxl.styles.fills import PatternFill
    from openpyxl.styles.fonts import Font as ExcelFont

    wb = Workbook()

    # ========================================
    # Sheet 1: 검사 결과 (이슈만)
    # ========================================
    ws_issues = wb.active
    ws_issues.title = "검사 결과"

    # 헤더
    headers = ['슬라이드', '원문', '수정안', '유형', '출처', '신뢰도', '이유']
    ws_issues.append(headers)

    # 헤더 스타일
    for col_num, header in enumerate(headers, 1):
        cell = ws_issues.cell(row=1, column=col_num)
        cell.font = ExcelFont(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 이슈 데이터
    for issue in report.get('issues', []):
        original = issue.get('original', '')
        suggested = issue.get('suggested', '')

        row_data = [
            issue.get('slide', ''),
            original,
            '',  # 수정안은 RichText로 별도 설정
            issue.get('type', ''),
            issue.get('source', ''),
            issue.get('confidence', ''),
            issue.get('reason', '')
        ]

        row_num = ws_issues.max_row + 1
        ws_issues.append(row_data)

        # 수정안 셀에 RichText 적용 (변경 부분만 빨간색)
        suggested_cell = ws_issues.cell(row=row_num, column=3)

        # Diff 계산
        matcher = difflib.SequenceMatcher(None, original, suggested)
        rich_text_parts = []

        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            text_part = suggested[j1:j2]
            if not text_part:  # 빈 문자열 스킵
                continue

            if tag == 'equal':
                # 동일한 부분은 검은색
                rich_text_parts.append(TextBlock(InlineFont(color="000000"), text_part))
            elif tag in ('replace', 'insert'):
                # 변경/추가된 부분은 빨간색
                rich_text_parts.append(TextBlock(InlineFont(color="FF0000"), text_part))
            # 'delete'는 수정안에 나타나지 않으므로 무시

        if rich_text_parts:
            suggested_cell.value = CellRichText(*rich_text_parts)
        else:
            suggested_cell.value = suggested

    # 열 너비 자동 조정
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        if header in ['원문', '수정안', '이유']:
            ws_issues.column_dimensions[col_letter].width = 50
        elif header == '유형':
            ws_issues.column_dimensions[col_letter].width = 12
        else:
            ws_issues.column_dimensions[col_letter].width = 15

    # ========================================
    # Sheet 2: 전체 텍스트 (문제 없음 포함)
    # ========================================
    if all_texts:
        ws_all = wb.create_sheet(title="전체 검사 현황")

        # 이슈가 있는 텍스트의 키 생성 (슬라이드 번호 + 원문)
        issue_keys = set()
        for issue in report.get('issues', []):
            key = (issue.get('slide'), issue.get('original', '').strip())
            issue_keys.add(key)

        # 헤더
        headers_all = ['슬라이드', '텍스트', '글자수', '검사 상태', '비고']
        ws_all.append(headers_all)

        # 헤더 스타일
        for col_num, header in enumerate(headers_all, 1):
            cell = ws_all.cell(row=1, column=col_num)
            cell.font = ExcelFont(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 슬라이드 번호로 정렬
        sorted_texts = sorted(all_texts, key=lambda x: x.get('slide', 0))

        # 모든 텍스트 데이터
        for text_item in sorted_texts:
            slide_num = text_item.get('slide', '')
            text = text_item.get('text', '').strip()
            char_count = text_item.get('char_count', len(text))

            # 이슈 여부 확인
            key = (slide_num, text)
            if key in issue_keys:
                status = "⚠️ 이슈 발견"
                note = "검사 결과 시트 참조"
                fill_color = "FFEEEE"  # 연한 빨강
            else:
                status = "✅ 문제 없음"
                note = ""
                fill_color = "EEFFEE"  # 연한 초록

            row_data = [slide_num, text, char_count, status, note]
            row_num = ws_all.max_row + 1
            ws_all.append(row_data)

            # 상태에 따라 배경색 적용
            for col_num in range(1, len(headers_all) + 1):
                cell = ws_all.cell(row=row_num, column=col_num)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # 열 너비 조정
        ws_all.column_dimensions['A'].width = 10  # 슬라이드
        ws_all.column_dimensions['B'].width = 60  # 텍스트
        ws_all.column_dimensions['C'].width = 10  # 글자수
        ws_all.column_dimensions['D'].width = 15  # 검사 상태
        ws_all.column_dimensions['E'].width = 20  # 비고

        print(f"   📋 전체 검사 현황: {len(sorted_texts)}개 텍스트")
        print(f"      ✅ 문제 없음: {len(sorted_texts) - len(issue_keys)}개")
        print(f"      ⚠️  이슈 발견: {len(issue_keys)}개")

    wb.save(excel_path)
    print(f"\n💾 Excel 리포트 저장: {excel_path}")
    print(f"   💡 수정안 열에서 변경된 부분만 빨간색으로 표시됩니다")
    print(f"   💡 '전체 검사 현황' 시트에서 모든 텍스트 확인 가능")


# ============================================
# 최종 리포트 생성
# ============================================

def generate_final_report(
    pptx_path: Path,
    texts: List[Dict],
    merged_issues: List[Dict],
    claude_result: Dict,
    output_path: Path = None
) -> Dict:
    """최종 리포트 생성"""
    print("\n" + "=" * 70)
    print("📊 최종 검사 리포트")
    print("=" * 70)

    # Claude 결과로 최종 이슈 리스트 업데이트
    final_issues = merged_issues.copy()

    # False positives 제거
    if claude_result.get('false_positives'):
        false_positive_keys = set()
        for fp in claude_result['false_positives']:
            false_positive_keys.add((fp.get('slide'), fp.get('original')))

        final_issues = [
            issue for issue in final_issues
            if (issue.get('slide'), issue['original']) not in false_positive_keys
        ]
        print(f"\n   False positives 제거: {len(merged_issues) - len(final_issues)}건")

    # Context-based 이슈 추가
    if claude_result.get('context_based_issues'):
        for issue in claude_result['context_based_issues']:
            final_issues.append({
                'source': 'claude_context',
                'slide': issue['slide'],
                'original': issue['original'],
                'suggested': issue['suggested'],
                'type': '문맥기반',
                'confidence': issue.get('confidence', 'high'),
                'reason': issue.get('reason', '')
            })

    # 추가 이슈 (reason 필드 포함)
    if claude_result.get('additional_issues'):
        for issue in claude_result['additional_issues']:
            final_issues.append({
                'source': 'claude_additional',
                'slide': issue['slide'],
                'original': issue.get('issue', ''),
                'suggested': issue.get('suggestion', ''),
                'type': '스타일/문맥',
                'confidence': 'medium',
                'reason': issue.get('reason', '')  # reason 필드 추가
            })

    # 타입별 분류
    by_type = defaultdict(list)
    for issue in final_issues:
        by_type[issue['type']].append(issue)

    # 파인튜닝 결과에 이유 생성
    final_issues = generate_reasons_for_finetuned(final_issues)

    # 원문과 수정안이 동일한 항목 제거
    print("\n🔍 동일한 원문/수정안 항목 제거 중...")
    original_count = len(final_issues)
    final_issues = [
        issue for issue in final_issues
        if issue.get('original', '').strip() != issue.get('suggested', '').strip()
    ]
    removed_count = original_count - len(final_issues)
    if removed_count > 0:
        print(f"   제거된 항목: {removed_count}건")
        print(f"   최종 이슈: {len(final_issues)}건")

    # 타입별 재분류 (중복 제거 후)
    by_type = defaultdict(list)
    for issue in final_issues:
        by_type[issue['type']].append(issue)

    # 리포트 생성
    report = {
        'file': str(pptx_path),
        'summary': {
            'total_texts': len(texts),
            'total_issues': len(final_issues),
            'by_type': {k: len(v) for k, v in by_type.items()},
            'by_source': {
                'finetuned_model': len([i for i in final_issues if i['source'] == 'finetuned_model']),
                'rule_based': len([i for i in final_issues if i['source'] == 'rule_based']),
                'claude_context': len([i for i in final_issues if i['source'] == 'claude_context']),
                'claude_additional': len([i for i in final_issues if i['source'] == 'claude_additional']),
            }
        },
        'issues': final_issues,
        'claude_review': claude_result
    }

    # 요약 출력
    print(f"\n📝 요약:")
    print(f"   - 검사된 텍스트: {len(texts)}개")
    print(f"   - 발견된 문제: {len(final_issues)}건")
    print(f"\n   타입별:")
    for issue_type, count in report['summary']['by_type'].items():
        print(f"      • {issue_type}: {count}건")
    print(f"\n   출처별:")
    for source, count in report['summary']['by_source'].items():
        if count > 0:
            print(f"      • {source}: {count}건")

    # 주요 이슈 출력
    if final_issues:
        print(f"\n🔍 주요 이슈 (최대 10개):")
        for i, issue in enumerate(final_issues[:10], 1):
            slide = issue.get('slide', '?')
            print(f"\n   {i}. [슬라이드 {slide}] {issue['type']}")
            print(f"      원문: \"{issue['original']}\"")
            print(f"      제안: \"{issue['suggested']}\"")
            if issue.get('reason'):
                print(f"      이유: {issue['reason']}")
            print(f"      (출처: {issue['source']}, 신뢰도: {issue['confidence']})")

        if len(final_issues) > 10:
            print(f"\n   ... 외 {len(final_issues) - 10}건")

    # JSON 저장
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        print(f"\n💾 리포트 저장: {output_path}")

    return report


# ============================================
# 메인 파이프라인
# ============================================

def run_integrated_check(pptx_path: str, output_path: str = None, model_path: str = None, skip_finetuned: bool = True):
    """통합 검사 파이프라인 실행

    Args:
        pptx_path: PPTX 파일 경로
        output_path: 출력 파일 경로 (JSON 또는 XLSX)
        model_path: 파인튜닝 모델 경로 (사용하지 않음, 하위 호환성 유지)
        skip_finetuned: 파인튜닝 모델 스킵 여부 (기본값: True, Knowledge Distillation 모드)
    """
    pptx_path = Path(pptx_path)

    if not pptx_path.exists():
        print(f"❌ 파일을 찾을 수 없습니다: {pptx_path}")
        return

    if model_path is None:
        model_path = "./qwen3-spelling-checker/final"

    # 출력 형식 자동 감지
    if output_path is None:
        output_path = pptx_path.parent / f"{pptx_path.stem}_integrated_report.json"

    output_path_str = str(output_path)

    if output_path_str.endswith('.xlsx'):
        output_format = 'excel'
        json_output_path = output_path_str.rsplit('.', 1)[0] + '.json'
    else:
        output_format = 'json'
        json_output_path = output_path_str

    print("\n" + "=" * 70)
    print("🔍 통합 PPTX 검사 파이프라인")
    if skip_finetuned:
        print("   💡 Knowledge Distillation 모드: Few-shot 예시로 파인튜닝 지식 활용")
    print("=" * 70)
    print(f"파일: {pptx_path.name}")
    print(f"출력: {output_path} ({output_format.upper()} 형식)")

    # Stage 1: 텍스트 추출
    texts = extract_texts_from_pptx(pptx_path)
    if not texts:
        print("❌ 텍스트 추출 실패")
        return

    # 추출된 텍스트를 중간 파일로 저장
    extracted_texts_path = pptx_path.parent / f"{pptx_path.stem}_extracted_texts.json"
    with open(extracted_texts_path, 'w', encoding='utf-8') as f:
        json.dump({
            'file': str(pptx_path),
            'total_texts': len(texts),
            'texts': texts
        }, f, ensure_ascii=False, indent=2)
    print(f"\n💾 추출된 텍스트 저장: {extracted_texts_path}")

    # Stage 2: 규칙 기반 검사 (파인튜닝 모델은 Knowledge Distillation으로 대체)
    print("\n" + "=" * 70)
    if skip_finetuned:
        print("🔬 Stage 2: 규칙 기반 검사")
        print("   ⏭️  파인튜닝 모델 스킵 (Knowledge Distillation 모드)")
    else:
        print("🔬 Stage 2: 파인튜닝 모델 + 규칙 기반 검사")
    print("=" * 70)

    if not skip_finetuned:
        model_issues = check_with_finetuned_model(texts, model_path)
    else:
        model_issues = []  # 파인튜닝 모델 스킵

    rule_issues = check_with_rules(pptx_path)
    merged_issues = merge_issues(model_issues, rule_issues)

    # Stage 3: Claude API 검토
    claude_result = review_with_claude(pptx_path, texts, merged_issues)

    # 최종 리포트 (JSON)
    report = generate_final_report(pptx_path, texts, merged_issues, claude_result, Path(json_output_path))

    # Excel 저장 (전체 텍스트 포함)
    if output_format == 'excel':
        save_report_as_excel(report, output_path_str, all_texts=texts)

    print("\n" + "=" * 70)
    print("✅ 통합 검사 완료!")
    print("=" * 70)

    return report


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("통합 PPTX 검사 파이프라인")
        print("-" * 40)
        print("사용법: python integrated_pptx_checker.py <input.pptx> [옵션]")
        print()
        print("옵션:")
        print("  --output PATH    결과 파일 경로 (.json 또는 .xlsx)")
        print("  --model PATH     파인튜닝 모델 경로 (기본: ./qwen3-spelling-checker/final)")
        print()
        print("예시:")
        print("  python integrated_pptx_checker.py presentation.pptx")
        print("  python integrated_pptx_checker.py presentation.pptx --output report.json")
        print("  python integrated_pptx_checker.py presentation.pptx --output report.xlsx")
        sys.exit(0)

    pptx_path = sys.argv[1]

    output_path = None
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            output_path = sys.argv[idx + 1]

    model_path = None
    if "--model" in sys.argv:
        idx = sys.argv.index("--model")
        if idx + 1 < len(sys.argv):
            model_path = sys.argv[idx + 1]

    run_integrated_check(pptx_path, output_path, model_path)
