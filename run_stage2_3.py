#!/usr/bin/env python3
"""
Stage 2-3 검사 실행 (기존 Vision 추출 결과 사용)
이미 추출된 test_pdf_vision_result.json을 사용하여 Stage 2-3만 실행

사용법:
    python run_stage2_3.py test_pdf_vision_result.json [--output report.csv]
"""

import json
import sys
import csv
import os
from pathlib import Path
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# integrated_pptx_checker의 Stage 2-3 함수들 임포트
from integrated_pptx_checker import (
    check_with_finetuned_model,
    check_with_rules,
    merge_issues,
    review_with_claude,
    generate_final_report
)


def load_vision_result(json_path: str):
    """Vision 추출 결과 JSON 로드"""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    print(f"📄 Vision 추출 결과 로드: {json_path}")
    print(f"   총 {len(data)}개 페이지")

    # integrated_pptx_checker 형식으로 변환
    texts = []
    for page_data in data:
        page_num = page_data.get('page', 0)
        for text in page_data.get('texts', []):
            if len(text) > 5:  # 최소 길이 필터
                texts.append({
                    'slide': page_num,
                    'text': text,
                    'char_count': len(text)
                })

    print(f"   변환된 텍스트: {len(texts)}개 블록")
    return texts


def highlight_diff(original: str, suggested: str) -> str:
    """
    원문과 수정안의 차이를 빨간색으로 표시
    [RED:텍스트] 형식으로 표시
    """
    import difflib

    # 완전히 동일하면 수정안 그대로 반환
    if original == suggested:
        return suggested

    # 문자 단위 diff
    matcher = difflib.SequenceMatcher(None, original, suggested)
    result = []

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'equal':
            result.append(suggested[j1:j2])
        elif tag == 'replace':
            result.append(f'[RED:{suggested[j1:j2]}]')
        elif tag == 'insert':
            result.append(f'[RED:{suggested[j1:j2]}]')
        elif tag == 'delete':
            # 삭제된 부분은 표시 안 함
            pass

    return ''.join(result)


def generate_reasons_for_finetuned(issues: list) -> list:
    """
    파인튜닝 모델 결과에 대해 Claude API로 이유 생성

    Args:
        issues: 이슈 리스트

    Returns:
        이유가 추가된 이슈 리스트
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("⚠️  ANTHROPIC_API_KEY 없음 - 파인튜닝 이유 생성 스킵")
        return issues

    # 파인튜닝 이슈만 필터
    finetuned_issues = [
        issue for issue in issues
        if issue.get('source') == 'finetuned_model' and not issue.get('reason')
    ]

    if not finetuned_issues:
        return issues

    print(f"\n🤖 Claude API로 파인튜닝 결과 이유 생성 중... ({len(finetuned_issues)}건)")

    client = anthropic.Anthropic(api_key=api_key)

    # 배치로 처리
    batch_data = []
    for issue in finetuned_issues:
        batch_data.append({
            "original": issue.get('original', ''),
            "suggested": issue.get('suggested', ''),
            "type": issue.get('type', '')
        })

    prompt = f"""다음은 한국어 맞춤법/용어 통일성 검사기가 발견한 수정 제안입니다.
각 수정 제안에 대해 **왜 수정이 필요한지** 간결하게 설명해주세요 (10-20자).

입력:
```json
{json.dumps(batch_data, ensure_ascii=False, indent=2)}
```

출력 형식 (JSON만):
```json
{{
  "reasons": [
    "설명1",
    "설명2",
    ...
  ]
}}
```

JSON만 출력하세요."""

    try:
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}]
        )

        response_text = message.content[0].text

        # JSON 추출
        json_str = response_text.strip()
        if '```json' in json_str:
            json_str = json_str.split('```json')[1].split('```')[0].strip()
        elif '```' in json_str:
            json_str = json_str.split('```')[1].split('```')[0].strip()

        data = json.loads(json_str)
        reasons = data.get('reasons', [])

        # 이슈에 이유 추가
        issue_map = {id(issue): issue for issue in issues}
        for i, finetuned_issue in enumerate(finetuned_issues):
            if i < len(reasons):
                issue_map[id(finetuned_issue)]['reason'] = reasons[i]

        print(f"   ✅ 이유 생성 완료")

    except Exception as e:
        print(f"   ⚠️  이유 생성 실패: {e}")

    return list(issue_map.values())


def save_report_as_excel(report: dict, excel_path: str):
    """리포트를 Excel 파일로 저장 (실제 빨간색 포맷팅)"""
    import difflib
    from openpyxl.styles.fills import PatternFill
    from openpyxl.styles.fonts import Font as ExcelFont

    wb = Workbook()
    ws = wb.active
    ws.title = "검사 결과"

    # 헤더
    headers = ['슬라이드', '원문', '수정안', '유형', '출처', '신뢰도', '이유']
    ws.append(headers)

    # 헤더 스타일
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = ExcelFont(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 이슈 데이터
    for issue in report.get('issues', []):
        original = issue.get('original', '')
        suggested = issue.get('suggested', '')

        row_data = [
            issue.get('slide', ''),
            original,
            suggested,  # 수정안은 diff 표시 없이
            issue.get('type', ''),
            issue.get('source', ''),
            issue.get('confidence', ''),
            issue.get('reason', '')
        ]

        row_num = ws.max_row + 1
        ws.append(row_data)

        # 수정안 셀에 빨간색 적용
        suggested_cell = ws.cell(row=row_num, column=3)

        # Diff 계산
        matcher = difflib.SequenceMatcher(None, original, suggested)
        has_diff = False

        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag in ('replace', 'insert'):
                has_diff = True
                break

        # 변경사항이 있으면 빨간색 폰트
        if has_diff:
            suggested_cell.font = ExcelFont(color="FF0000")  # 빨간색

    # 열 너비 자동 조정
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        if header in ['원문', '수정안', '이유']:
            ws.column_dimensions[col_letter].width = 50
        elif header == '유형':
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 15

    wb.save(excel_path)
    print(f"\n💾 Excel 리포트 저장: {excel_path}")
    print(f"   💡 수정안 열에서 변경된 항목은 빨간색으로 표시됩니다")


def save_report_as_csv(report: dict, csv_path: str):
    """리포트를 CSV 파일로 저장 (변경 부분 빨간색 표시)"""
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)

        # 헤더
        writer.writerow([
            '슬라이드',
            '원문',
            '수정안 (변경부분 [RED:...] 표시)',
            '유형',
            '출처',
            '신뢰도',
            '이유'
        ])

        # 이슈 데이터
        for issue in report.get('issues', []):
            original = issue.get('original', '')
            suggested = issue.get('suggested', '')

            # 변경 부분 하이라이트
            highlighted = highlight_diff(original, suggested)

            writer.writerow([
                issue.get('slide', ''),
                original,
                highlighted,
                issue.get('type', ''),
                issue.get('source', ''),
                issue.get('confidence', ''),
                issue.get('reason', '')
            ])

    print(f"\n💾 CSV 리포트 저장: {csv_path}")
    print(f"   💡 [RED:...] 표시는 변경된 부분입니다")


def run_stage2_3(json_path: str, output_path: str = None, model_path: str = None):
    """Stage 2-3 검사 실행"""

    if model_path is None:
        model_path = "./qwen3-spelling-checker/final"

    if output_path is None:
        output_path = Path(json_path).stem + "_stage2_3_report.csv"

    # 출력 형식 자동 감지 (확장자 기반)
    if output_path.endswith('.csv'):
        output_format = 'csv'
    elif output_path.endswith('.xlsx'):
        output_format = 'excel'
    else:
        output_format = 'json'

    # JSON 출력 경로도 유지 (내부 처리용)
    if output_format in ('csv', 'excel'):
        json_output_path = output_path.rsplit('.', 1)[0] + '.json'
    else:
        json_output_path = output_path

    print("\n" + "=" * 70)
    print("🔍 Stage 2-3 검사 시작")
    print("=" * 70)
    print(f"입력: {json_path}")
    print(f"출력: {output_path} ({output_format.upper()} 형식)")
    print(f"모델: {model_path}")

    # Stage 1: 기존 Vision 결과 로드
    print("\n" + "=" * 70)
    print("📥 Stage 1: 기존 Vision 추출 결과 로드")
    print("=" * 70)
    texts = load_vision_result(json_path)

    if not texts:
        print("❌ 텍스트가 없습니다")
        return

    # Stage 2: 파인튜닝 + 규칙 기반
    print("\n" + "=" * 70)
    print("🔬 Stage 2: 파인튜닝 모델 + 규칙 기반 검사")
    print("=" * 70)

    model_issues = check_with_finetuned_model(texts, model_path)

    # pptx_path 없이는 규칙 기반 검사 불가 (스킵)
    print("\n⚠️  규칙 기반 검사 스킵 (PPTX 파일 없음)")
    rule_issues = []

    merged_issues = merge_issues(model_issues, rule_issues)

    # Stage 3: Claude API 검토
    print("\n" + "=" * 70)
    print("🤖 Stage 3: Claude API 최종 검토")
    print("=" * 70)

    claude_result = review_with_claude(None, texts, merged_issues)

    # 최종 리포트 생성 (JSON)
    report = generate_final_report(
        Path(json_path),
        texts,
        merged_issues,
        claude_result,
        Path(json_output_path)
    )

    # 파인튜닝 결과에 이유 생성
    report['issues'] = generate_reasons_for_finetuned(report['issues'])

    # 원문과 수정안이 동일한 항목 제거
    print("\n🔍 동일한 원문/수정안 항목 제거 중...")
    original_count = len(report['issues'])
    report['issues'] = [
        issue for issue in report['issues']
        if issue.get('original', '').strip() != issue.get('suggested', '').strip()
    ]
    removed_count = original_count - len(report['issues'])
    if removed_count > 0:
        print(f"   제거된 항목: {removed_count}건")
        print(f"   최종 이슈: {len(report['issues'])}건")

    # JSON 다시 저장 (이유 추가 + 동일 항목 제거 후)
    with open(json_output_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    # CSV 또는 Excel 저장
    if output_format == 'csv':
        save_report_as_csv(report, output_path)
    elif output_format == 'excel':
        save_report_as_excel(report, output_path)

    print("\n" + "=" * 70)
    print("✅ Stage 2-3 검사 완료!")
    print("=" * 70)

    return report


def main():
    if len(sys.argv) < 2:
        print("Stage 2-3 검사 실행")
        print("-" * 40)
        print("사용법: python run_stage2_3.py <vision_result.json> [옵션]")
        print()
        print("옵션:")
        print("  --output PATH    결과 파일 경로 (.csv 또는 .json)")
        print("  --model PATH     파인튜닝 모델 경로 (기본: ./qwen3-spelling-checker/final)")
        print()
        print("예시:")
        print("  python run_stage2_3.py test_pdf_vision_result.json")
        print("  python run_stage2_3.py test_pdf_vision_result.json --output report.csv")
        print("  python run_stage2_3.py test_pdf_vision_result.json --output report.json")
        sys.exit(0)

    json_path = sys.argv[1]

    output_path = None
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            output_path = sys.argv[idx + 1]

    model_path = None
    if "--model" in sys.argv:
        idx = sys.argv.index("--model")
        if idx + 1 < len(sys.argv):
            model_path = sys.argv[idx + 1]

    run_stage2_3(json_path, output_path, model_path)


if __name__ == "__main__":
    main()